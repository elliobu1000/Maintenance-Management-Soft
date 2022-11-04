
import enum
from PyQt5 import QtCore, QtWidgets, uic
from PyQt5.QtWidgets import QMessageBox
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import pandas as pd
import datetime as dt
import calendar
import sys

class Ui(QtWidgets.QMainWindow):
    def __init__(self):
        super(Ui, self).__init__()
        uic.loadUi('gui/Gui.ui', self)

        self.addButton.clicked.connect(self.alpha)
        self.addButton.setShortcut("Return")
        self.deleteButton.clicked.connect(self.beta)
        self.mtbfButton.clicked.connect(self.calendar)
        self.update_excel()
        self.show()



    def update_excel(self):

        wb = load_workbook('file.xlsx')
        ws = wb.active
        
        for i in range(2,len(ws['E'])):
            ws['E'+str(i)].value = str(ws['E'+str(i)].value).capitalize()
        wb.save('file.xlsx')

    def popup(self, title = 'Title', text = 'Text', info = 'info', flag = QMessageBox.Information):
        #QMessageBox.Question
        #QMessageBox.Information
        #QMessageBox.Warning
        #QMessageBox.Critical
        msg = QMessageBox()
        msg.setIcon(flag)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setInformativeText(info)
        msg.setStandardButtons(QMessageBox.Ok)
        msg.exec_()

    def check(self, list, labels, wb):
        txt = ''
        t = [n for i, n in zip(list, labels) if i == '']
        txt = '\n'.join(t)
        if txt == '':
            self.popup(text = 'Les Informations Sont Enregitrées avec Succées avec ', info = f"NO ° : {list[0]}")
            wb.save('file.xlsx')
        else :
            self.popup(text = txt, info = 'Vide!', flag = QMessageBox.Critical)


    def alpha(self):
        wb = load_workbook('file.xlsx')
        ws = wb['Feuil1']
        self.i = len(ws['D'])

        try :
            L = int(str(self.lineEdit.text()))
        except ValueError:
            L = 0

        list_of_line_edits = [
            L, str(self.lineEdit_2.text()), self.comboBox.currentText(), str(self.lineEdit_3.text()), str(self.lineEdit_6.text()),
            str(self.lineEdit_5.text()), str(self.timeEdit_5.text()), str(self.timeEdit_2.text()), str(self.timeEdit.text()),
            str(self.timeEdit_4.text()), str(self.lineEdit_4.text())
            ]
        d = dt.datetime.strptime('0'+list_of_line_edits[-2].split(' ')[0]+':00', '%H:%M:%S')

        if list_of_line_edits[0] in list([ws['B'+str(i+1)].value for i in range(1,len(ws['B']))]):
            self.popup(text = 'Attention!', info = f'Le N°OT : {list_of_line_edits[0]} existe déjà', flag = QMessageBox.Warning)
            return
        
        if d > dt.datetime(1900,1,1,0,5,0) and list_of_line_edits[-1] == '':
            self.popup(text = 'Attention!', info = 'Donné Une Justification Pour TR > 5 minute!', flag = QMessageBox.Warning)
            return

        ws['A'+str(self.i + 1)] = dt.date.today().strftime("%d/%m/%Y")
        for i, n in enumerate(list_of_line_edits):
            ws[chr(65 + i + 1) + str(self.i+1)] = n

        heure_fin_dintervention = dt.datetime(1990,1,1,int(str(self.timeEdit.text()).split(':')[0]), int(str(self.timeEdit.text()).split(':')[1]),0)
        heure_arret_machine = dt.timedelta(hours = int(str(self.timeEdit_5.text()).split(':')[0]), minutes = int(str(self.timeEdit_5.text()).split(':')[1]))
        d = heure_fin_dintervention - heure_arret_machine
        # temps_arret_machine = [int(heure_fin_dintervention[0]) - int(heure_arret_machine[0]), int(heure_fin_dintervention[1]) - int(heure_arret_machine[1])]
        ws['M'+ str(self.i + 1)] = f'{d.hour}:{d.minute}0'

        line_edits = [str(self.lineEdit.text()), str(self.lineEdit_2.text()), str(self.lineEdit_3.text()), str(self.lineEdit_6.text()), str(self.lineEdit_5.text())]
        labels = [self.label.text(), self.label_3.text(), self.label_2.text(), self.label_4.text(), self.label_10.text()]

        self.check(line_edits, labels, wb)

    def to_datetime(self, dates):
        for n, i in enumerate(dates):
            if type(i) == str:
                dates[n] = dt.datetime(int(i.split('/')[2]),int(i.split('/')[0]),int(i.split('/')[1]))
        return dates
    def to_time(self, time):
        for n, i in enumerate(time):
            if type(i) == str:
                time[n] = dt.datetime(1990,1,1,int(i.split(':')[0]),int(i.split(':')[1]),0)
        return time

    def calendar(self):
        wb = load_workbook('file.xlsx')
        ws = wb['Feuil1']

        mois = ['Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin', 'Juillet', 'Aôut', 'Septembre', 'Octobre', 'Novembre', 'Décembre']
        dates = self.to_datetime([ws['A'+str(i + 2)].value for i in range(len(ws['A'])) if ws['A'+str(i + 2)].value != None])
#        print(dates)
        # dates = [dt.datetime(int(i.split('/')[2]),int(i.split('/')[0]),int(i.split('/')[1])) for i in dates if type(i) == str]
        # dates = self.to_datetime(dates)
        reactivite = [ws['K'+str(i + 2)].value for i in range(len(ws['K'])) if ws['K'+str(i + 2)].value != None]
        arret_machine = [ws['M'+str(i + 2)].value for i in range(len(ws['M'])) if ws['M'+str(i + 2)].value != None]
        equipements = list(set([ws['E'+str(i+1)].value for i in range(1,len(ws['E']))]))
        equipements.remove(None)
        equip = list([ws['E'+str(i+1)].value for i in range(1,len(ws['E'])) if ws['E'+str(i + 1)].value != None])

        annee_souhaite = int(self.comboBox_2.currentText())
        mois_souhaite = mois.index(self.comboBox_3.currentText()) + 1


        print(mois_souhaite)
        nombre_de_jours_dans_mois = calendar.monthrange(annee_souhaite, mois_souhaite)[1]
        dates_de_pannes = [n for i, n in enumerate(dates) if mois_souhaite == dates[i].month and annee_souhaite == dates[i].year]
        #print(dates_de_pannes)

        # print(dates_de_pannes)

        for i, n in enumerate(dates):
            if n in dates_de_pannes:
                pass
                # print(i, equip[i])

        # idx = [dates.index(i) for i in dates_de_pannes]
        # [print(i, dates[i]) for i in idx]
        # [print(i, n) for i, n in enumerate(dates)]
        machines_en_panne = [equip[i] for i, n in enumerate(dates) if n in dates_de_pannes]
        n_machines_en_panne = [machines_en_panne.count(t) for t in equipements]

        # heure_rectivite = [round(i.hour + i.minute/60, 3) for i, j, k in zip(reactivite, machines_en_panne, equip) if j == k]
        # somme_heure_rectivite = [sum([i for i, j in zip(heure_rectivite, equip) if j == k]) for k in equipements]

        heure_arret_machine = [round(arret_machine[i].hour + arret_machine[i].minute/60, 3) for i, n in enumerate(equip) if n in machines_en_panne]
        somme_heure_arret_machine = [sum([i for i, j in zip(heure_arret_machine, equip) if j == k]) for k in equipements]

        # heure_arret_machine = [i for i, j, k in zip(arret_machine, dates, dates_de_pannes) if j == k]
        # somme_heure_rectivite = round(sum([i.hour for i in heure_rectivite]) + (sum([i.minute for i in heure_rectivite]))/60, 2)
        # somme_heure_arret_machine = round(sum([i.hour for i in heure_arret_machine]) + (sum([i.minute for i in heure_arret_machine]))/60, 2)

        # Si une machine est insérer 1 fois par mois
        for i, n in enumerate(n_machines_en_panne):
            if n == 0:
                n_machines_en_panne[i] += 1

        MTBF = [(nombre_de_jours_dans_mois)/(i) for i in n_machines_en_panne]
        MTTR = [(somme_heure_arret_machine[i])/(n) for i, n in enumerate(n_machines_en_panne)]
        TD = [(1 - (i)/(nombre_de_jours_dans_mois*24))*100 for i in somme_heure_arret_machine]

        for i, n in enumerate(MTBF):
            if n < 0:
                MTBF[i] = 0
        for i, n in enumerate(MTTR):
            if n < 0:
                MTTR[i] = 0

        # Titre = ', '.join(equipements)
        D = f'KPI du mois {mois[mois_souhaite-1]} {annee_souhaite}'
        self.plot(equipements, MTBF, MTTR, TD, D)


    def plot(self, equipements, MTBF = 0, MTTR = 0, TD = 0, D = 0):
        plt.subplot(3,1,1)
        plt.bar(equipements, MTBF)
        plt.xlabel('Les Machines')
        plt.ylabel('MTBF en Jour')
        plt.title(D)
        plt.subplot(3,1,2)
        plt.bar(equipements, MTTR)
        plt.xlabel('Les Machines')
        plt.ylabel('MTTR en Heure')

        plt.subplot(3,1,3)
        plt.bar(equipements, TD)
        plt.xlabel('Les Machines')
        plt.ylabel('TD en Pourcentage')

        plt.show()

    def beta(self):
        wb = load_workbook('file.xlsx')
        ws = wb['Feuil1']

        if self.lineEdit_11.text() == '':
            self.popup(text = "Warning : ", info = "Inserer un Nombre!", flag = QMessageBox.Warning)
            return

        f = list([ws['B'+str(i+1)].value for i in range(1,len(ws['B']))])
        x = f.index(int(self.lineEdit_11.text())) + 2

        for i in range(len(ws['1'])):
            ws[chr(65 + i) + str(x)] = None
        
        f = [[ws[chr(65 + i) + str(j)].value for i in range(len(ws['1']))] for j in range(x + 1, len(ws['B']) + 1)]

        for j in range(x, x + len(f)):
            for i in range(len(ws['1'])):
                ws[chr(65 + i) + str(j)] = f[j - x][i]

        for i in range(len(ws['1'])):
            ws[chr(65 + i) + str(len(ws['B']))] = None

        self.popup(info = "Infos", text = f"NO° : {int(self.lineEdit_11.text())} a été supprimer")

        wb.save('file.xlsx')

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = Ui()
    app.exec_()